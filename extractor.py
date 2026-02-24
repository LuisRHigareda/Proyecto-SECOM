import openpyxl
from openpyxl.utils import get_column_letter

# --- 1. CONFIGURACIÓN ---
# ¡IMPORTANTE! Asegúrate de que este sea el nombre exacto de tu archivo.
excel_filename = "COTIZACION SISTEMA FOTOVOLTAICO BIMESTRAL.xlsm"
output_filename = "formulas.txt"
# ------------------------

print(f"Abriendo el archivo: {excel_filename}...")

# Carga el libro de Excel, asegurándote de leer las fórmulas (data_only=False)
try:
    # Debes tener el archivo .xlsm en la misma carpeta que este script
    wb = openpyxl.load_workbook(excel_filename, data_only=False, keep_vba=True)
except FileNotFoundError:
    print(f"Error: No se encontró el archivo '{excel_filename}' en esta carpeta.")
    input("Presiona Enter para salir.")
    exit()
except Exception as e:
    print(f"Ocurrió un error al abrir el archivo: {e}")
    input("Presiona Enter para salir.")
    exit()

# Abre el archivo de texto donde guardaremos las fórmulas
with open(output_filename, "w", encoding="utf-8") as f:
    f.write(f"--- Fórmulas encontradas en: {excel_filename} ---\n\n")
    print(f"Extrayendo fórmulas en {output_filename}...")

    # Itera sobre cada hoja en el libro
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        formulas_encontradas = False
        
        # Itera sobre cada celda en la hoja
        for row in sheet.iter_rows():
            for cell in row:
                # Revisa si la celda tiene una fórmula (empieza con '=')
                if cell.data_type == 'f':
                    if not formulas_encontradas:
                        # Si es la primera fórmula en la hoja, escribe el nombre de la hoja
                        f.write(f"\n=== Hoja: {sheet_name} ===\n")
                        formulas_encontradas = True
                        
                    # Escribe la celda y su fórmula en el archivo de texto
                    f.write(f"Celda {cell.coordinate}: {cell.value}\n")

print(f"\n¡Listo! Todas las fórmulas se han guardado en el archivo: {output_filename}")
print("Por favor, sube o copia el contenido de 'formulas.txt' en nuestra conversación.")
input("\nPresiona Enter para cerrar esta ventana.")